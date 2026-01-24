"""
720システム出力生成
720システム用のExcelファイルを作成
"""

import logging
from typing import List, Optional
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from config import SYSTEM_720
from exceptions import OutputError
from utils import safe_str, safe_int
from models import System720Row

logger = logging.getLogger(__name__)


class System720Generator:
    """720システム出力生成クラス"""

    def __init__(
            self,
            frame_part_number: str,
            df_cm_picking: pd.DataFrame,
            df_a_parts_picking: Optional[pd.DataFrame],
            df_matrix: pd.DataFrame
    ):
        """
        初期化

        Args:
            frame_part_number: フレーム品番
            df_cm_picking: CMピッキングDataFrame
            df_a_parts_picking: A部品ピッキングDataFrame（Noneの場合はCラインモード）
            df_matrix: 構成表マトリックスDataFrame
        """
        self.frame_part_number = frame_part_number
        self.df_cm_picking = df_cm_picking
        self.df_a_parts_picking = df_a_parts_picking
        self.df_matrix = df_matrix
        self.slip_number = self._generate_slip_number()
        self.delivery_date = self._calculate_delivery_date()

    def generate(self) -> Workbook:
        """
        720システム入力用Excelを生成

        Returns:
            Workbook

        Raises:
            OutputError: 出力エラー
        """
        try:
            logger.info("720システム出力生成中...")

            system_rows = []
            line_number = 1

            # CMピッキングから追加
            for _, row in self.df_cm_picking.iterrows():
                system_row = System720Row(
                    slip_number=self.slip_number,
                    line_number=line_number,
                    part_number=safe_str(row['部品番号']),
                    quantity=safe_int(row['数量']),
                    unit='個',
                    delivery_date=self.delivery_date,
                    remarks=safe_str(row['オプション'])
                )
                system_rows.append(system_row)
                line_number += 1

            # A部品ピッキングから追加
            if self.df_a_parts_picking is not None and not self.df_a_parts_picking.empty:
                for _, row in self.df_a_parts_picking.iterrows():
                    system_row = System720Row(
                        slip_number=self.slip_number,
                        line_number=line_number,
                        part_number=safe_str(row['部品番号']),
                        quantity=safe_int(row['数量']),
                        unit='個',
                        delivery_date=self.delivery_date,
                        remarks=safe_str(row['オプション'])
                    )
                    system_rows.append(system_row)
                    line_number += 1

            wb = self._create_workbook(system_rows)

            logger.info(f"720システム出力生成完了: {len(system_rows)}行")
            return wb

        except Exception as e:
            raise OutputError(f"720システム出力生成エラー: {e}")

    def _generate_slip_number(self) -> str:
        """
        伝票番号を生成

        Returns:
            伝票番号
        """
        now = datetime.now()
        date_part = now.strftime('%Y%m%d')
        seq_part = '001'

        return f"{date_part}-{seq_part}"

    def _calculate_delivery_date(self, days_ahead: int = 7) -> str:
        """
        納期を計算

        Args:
            days_ahead: 何日後（デフォルト: 7日）

        Returns:
            納期（YYYY-MM-DD形式）
        """
        delivery = datetime.now() + timedelta(days=days_ahead)
        return delivery.strftime('%Y-%m-%d')

    def _create_workbook(self, system_rows: List[System720Row]) -> Workbook:
        """
        Workbookを作成

        Args:
            system_rows: 720システム行のリスト

        Returns:
            Workbook
        """
        wb = Workbook()
        ws = wb.active
        ws.title = SYSTEM_720['sheet_name']

        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー行
        headers = SYSTEM_720['columns']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # データ行
        for row_idx, system_row in enumerate(system_rows, start=2):
            data = system_row.to_list()
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = border

        # 列幅調整
        column_widths = {
            1: 15,  # 伝票番号
            2: 8,  # 行番号
            3: 20,  # 品番
            4: 8,  # 数量
            5: 6,  # 単位
            6: 12,  # 納期
            7: 30  # 備考
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[chr(64 + col_idx)].width = width

        return wb